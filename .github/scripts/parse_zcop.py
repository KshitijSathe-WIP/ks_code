#!/usr/bin/env python3
"""
ZCOP Excel Parser - Extract key metrics from daily ZCOP files.

Usage:
  python parse_zcop.py <xlsx_file_path> [--save-snapshot]
  
Extracts and prints metrics for:
  - Billable/Non-billable hours and costs
  - NGA allocations
  - RD counts by type
  - Metrics by DM, business line, and client VP
"""

import json
import sys
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Dict, List, Any, Optional

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)


class ZCOPParser:
    """Parse ZCOP Excel files and extract key metrics."""
    
    # Common column name patterns
    PATTERNS = {
        'billable': r'billable|bill.*hours|billed',
        'non_billable': r'non.billable|non.bill|unb|unbill',
        'nga': r'nga|nearshore|academy',
        'ramp_up': r'ramp.*up|ru\b|rampup|ram#p.*up',
        'ramp_down': r'ramp.*down|rd\b|rampdown|ramp.*dn',
        'dm': r'dm\b|delivery.*manager|mgr',
        'business_line': r'business.*line|bl\b|practice',
        'client_vp': r'client.*vp|vp\b|account|VP',
        'cost': r'cost|amount|revenue|rate',
        'hours': r'hours|hrs|time'
    }
    
    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        self.wb = None
        self.ws = None
        self.headers = []
        self.data = []
        self.metrics = {
            'date': self.extract_date_from_filename(),
            'file_source': self.filepath.name,
            'metrics': {
                'billable': {'hours': 0, 'cost': 0},
                'non_billable': {'hours': 0, 'cost': 0},
                'nga': {'allocation_count': 0, 'percentage_of_total': 0},
                'ramp_up': {'count': 0, 'percentage_of_billable': 0},
                'ramp_down': {'count': 0, 'percentage_of_billable': 0}
            },
            'dimensions': {
                'by_dm': {},
                'by_business_line': {},
                'by_client_vp': {}
            },
            'anomalies': [],
            'notes': ''
        }
        
    def extract_date_from_filename(self) -> str:
        """Extract date from filename like 'TD Bank ZCOP 28-Jan-2026.xlsx'"""
        match = re.search(r'(\d{1,2})-(\w+)-(\d{4})', self.filepath.name)
        if match:
            day, month_str, year = match.groups()
            try:
                date_obj = datetime.strptime(f"{day}-{month_str}-{year}", "%d-%b-%Y")
                return date_obj.strftime("%Y-%m-%d")
            except ValueError:
                pass
        return datetime.now().strftime("%Y-%m-%d")
    
    def load_file(self) -> bool:
        """Load and read Excel file."""
        try:
            self.wb = openpyxl.load_workbook(self.filepath, data_only=True)
            # Use first sheet
            self.ws = self.wb.active
            
            # Extract headers from first row
            self.headers = [cell.value for cell in self.ws[1]]
            
            # Extract data rows
            for row_idx in range(2, self.ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(self.headers, 1):
                    if header:
                        cell = self.ws.cell(row_idx, col_idx)
                        row_data[header] = cell.value
                self.data.append(row_data)
            
            return True
        except Exception as e:
            print(f"ERROR: Failed to load {self.filepath}: {e}")
            return False
        finally:
            if self.wb:
                self.wb.close()
    
    def find_column_by_pattern(self, pattern: str) -> Optional[str]:
        """Find header column matching regex pattern."""
        regex = re.compile(pattern, re.IGNORECASE)
        for header in self.headers:
            if header and regex.search(str(header)):
                return header
        return None
    
    def parse_metrics(self) -> Dict[str, Any]:
        """Parse and aggregate metrics from data."""
        billable_col = self.find_column_by_pattern(self.PATTERNS['billable'])
        non_billable_col = self.find_column_by_pattern(self.PATTERNS['non_billable'])
        nga_col = self.find_column_by_pattern(self.PATTERNS['nga'])
        ramp_up_col = self.find_column_by_pattern(self.PATTERNS['ramp_up'])
        ramp_down_col = self.find_column_by_pattern(self.PATTERNS['ramp_down'])
        dm_col = self.find_column_by_pattern(self.PATTERNS['dm'])
        bl_col = self.find_column_by_pattern(self.PATTERNS['business_line'])
        client_vp_col = self.find_column_by_pattern(self.PATTERNS['client_vp'])
        cost_col = self.find_column_by_pattern(self.PATTERNS['cost'])
        
        # Initialize dimension tracking
        dm_counts = defaultdict(lambda: {'count': 0, 'billable_hours': 0, 'non_billable_hours': 0})
        bl_counts = defaultdict(lambda: {'count': 0, 'billable_hours': 0})
        cv_counts = defaultdict(lambda: {'count': 0, 'billable_hours': 0})
        ramp_up_count = 0
        ramp_down_count = 0
        nga_count = 0
        total_billable_tracked = 0
        
        # Process each data row
        for row in self.data:
            try:
                # Billable metrics
                if billable_col and row.get(billable_col):
                    val = float(row[billable_col]) if isinstance(row[billable_col], (int, float)) else 0
                    self.metrics['metrics']['billable']['hours'] += val
                
                # Non-billable metrics
                if non_billable_col and row.get(non_billable_col):
                    val = float(row[non_billable_col]) if isinstance(row[non_billable_col], (int, float)) else 0
                    self.metrics['metrics']['non_billable']['hours'] += val
                
                # NGA tracking
                if nga_col and row.get(nga_col) and str(row[nga_col]).strip().upper() in ['Y', 'YES', '1']:
                    nga_count += 1
                
                # Ramp Up tracking
                if ramp_up_col and row.get(ramp_up_col) and str(row[ramp_up_col]).strip().upper() in ['Y', 'YES', '1', 'TRUE']:
                    ramp_up_count += 1
                
                # Ramp Down tracking
                if ramp_down_col and row.get(ramp_down_col) and str(row[ramp_down_col]).strip().upper() in ['Y', 'YES', '1', 'TRUE']:
                    ramp_down_count += 1
                
                # Dimension tracking
                billable_hours = float(row.get(billable_col, 0)) if billable_col else 0
                non_billable_hours = float(row.get(non_billable_col, 0)) if non_billable_col else 0
                
                if dm_col and row.get(dm_col):
                    dm = str(row[dm_col]).strip()
                    dm_counts[dm]['count'] += 1
                    dm_counts[dm]['billable_hours'] += billable_hours
                    dm_counts[dm]['non_billable_hours'] += non_billable_hours
                
                if bl_col and row.get(bl_col):
                    bl = str(row[bl_col]).strip()
                    bl_counts[bl]['count'] += 1
                    bl_counts[bl]['billable_hours'] += billable_hours
                
                if client_vp_col and row.get(client_vp_col):
                    cv = str(row[client_vp_col]).strip()
                    cv_counts[cv]['count'] += 1
                    cv_counts[cv]['billable_hours'] += billable_hours
                
            except (ValueError, TypeError):
                continue
        
        # Store aggregated metrics
        self.metrics['metrics']['nga']['allocation_count'] = nga_count
        total_rows = len(self.data) if self.data else 1
        self.metrics['metrics']['nga']['percentage_of_total'] = (nga_count / total_rows * 100) if total_rows > 0 else 0
        
        self.metrics['metrics']['ramp_up']['count'] = ramp_up_count
        self.metrics['metrics']['ramp_up']['percentage_of_billable'] = (ramp_up_count / total_rows * 100) if total_rows > 0 else 0
        
        self.metrics['metrics']['ramp_down']['count'] = ramp_down_count
        self.metrics['metrics']['ramp_down']['percentage_of_billable'] = (ramp_down_count / total_rows * 100) if total_rows > 0 else 0
        
        self.metrics['dimensions']['by_dm'] = {k: dict(v) for k, v in sorted(dm_counts.items())}
        self.metrics['dimensions']['by_business_line'] = {k: dict(v) for k, v in sorted(bl_counts.items())}
        self.metrics['dimensions']['by_client_vp'] = {k: dict(v) for k, v in sorted(cv_counts.items())}
        
        # Detect anomalies
        if self.metrics['metrics']['billable']['hours'] == 0 and self.metrics['metrics']['non_billable']['hours'] == 0:
            self.metrics['anomalies'].append("No billable or non-billable hours found")
        
        if nga_count > total_rows * 0.5:
            self.metrics['anomalies'].append(f"High NGA allocation: {self.metrics['metrics']['nga']['percentage_of_total']:.1f}%")
        
        return self.metrics
    
    def update_snapshot_file(self, snapshot_file: str = None) -> bool:
        """Update the central metrics tracking file with this snapshot."""
        if not snapshot_file:
            # Path from ZCOP file to .github/data/zcop-metrics.json
            # Zcop Analysis/file.xlsx -> .github/data/zcop-metrics.json
            snapshot_file = self.filepath.parent.parent / '.github' / 'data' / 'zcop-metrics.json'
        
        try:
            # Load existing snapshots
            if Path(snapshot_file).exists():
                with open(snapshot_file, 'r') as f:
                    data = json.load(f)
            else:
                data = {
                    'schema_version': '1.0',
                    'last_updated': None,
                    'snapshots': [],
                    'aggregations': {
                        'billable_total': 0,
                        'non_billable_total': 0,
                        'nga_total': 0,
                        'ramp_up_total': 0,
                        'ramp_down_total': 0
                    }
                }
            
            # Remove duplicate entries for this date
            data['snapshots'] = [s for s in data['snapshots'] if s.get('date') != self.metrics['date']]
            
            # Add new snapshot
            data['snapshots'].append(self.metrics)
            data['last_updated'] = datetime.now().isoformat()
            
            # Update aggregations (sum of all snapshots)
            billable_total = sum(s['metrics']['billable']['hours'] for s in data['snapshots'])
            non_billable_total = sum(s['metrics']['non_billable']['hours'] for s in data['snapshots'])
            nga_total = sum(s['metrics']['nga']['allocation_count'] for s in data['snapshots'])
            ramp_up_total = sum(s['metrics']['ramp_up']['count'] for s in data['snapshots'])
            ramp_down_total = sum(s['metrics']['ramp_down']['count'] for s in data['snapshots'])
            
            data['aggregations'] = {
                'billable_total': billable_total,
                'non_billable_total': non_billable_total,
                'nga_total': nga_total,
                'ramp_up_total': ramp_up_total,
                'ramp_down_total': ramp_down_total
            }
            
            # Sort snapshots by date descending
            data['snapshots'].sort(key=lambda x: x['date'], reverse=True)
            
            # Write updated file
            with open(snapshot_file, 'w') as f:
                json.dump(data, f, indent=2)
            
            return True
        except Exception as e:
            print(f"ERROR: Failed to update snapshot file: {e}")
            return False
    
    def print_summary(self):
        """Print formatted summary of metrics."""
        m = self.metrics['metrics']
        print(f"\n=== ZCOP Analysis Summary ({self.metrics['date']}) ===")
        print(f"Source: {self.metrics['file_source']}\n")
        
        print("Key Metrics:")
        print(f"  Billable Hours: {m['billable']['hours']:.1f}")
        print(f"  Non-Billable Hours: {m['non_billable']['hours']:.1f}")
        print(f"  NGA Allocations: {m['nga']['allocation_count']} ({m['nga']['percentage_of_total']:.1f}%)")
        print(f"  Ramp Up (RU): {m['ramp_up']['count']} ({m['ramp_up']['percentage_of_billable']:.1f}% of workforce)")
        print(f"  Ramp Down (RD): {m['ramp_down']['count']} ({m['ramp_down']['percentage_of_billable']:.1f}% of workforce)")
        
        print("\nBy Dimension:")
        if self.metrics['dimensions']['by_dm']:
            print(f"  Top DMs: {list(self.metrics['dimensions']['by_dm'].keys())[:3]}")
        if self.metrics['dimensions']['by_business_line']:
            print(f"  Business Lines: {list(self.metrics['dimensions']['by_business_line'].keys())[:3]}")
        if self.metrics['dimensions']['by_client_vp']:
            print(f"  Client VPs: {list(self.metrics['dimensions']['by_client_vp'].keys())[:3]}")
        
        if self.metrics['anomalies']:
            print(f"\nAnomalies: {self.metrics['anomalies']}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python parse_zcop.py <xlsx_file_path> [--save-snapshot]")
        sys.exit(1)
    
    filepath = sys.argv[1]
    save_snapshot = '--save-snapshot' in sys.argv
    
    parser = ZCOPParser(filepath)
    
    if not parser.load_file():
        sys.exit(1)
    
    parser.parse_metrics()
    parser.print_summary()
    
    if save_snapshot:
        if parser.update_snapshot_file():
            print("\n✓ Snapshot saved to metrics tracking file")
        else:
            print("\n✗ Failed to save snapshot")
    
    # Output metrics as JSON for programmatic use
    print("\n=== JSON Output ===")
    print(json.dumps(parser.metrics, indent=2))


if __name__ == '__main__':
    main()
