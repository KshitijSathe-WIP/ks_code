#!/usr/bin/env python3
"""
ZCOP Excel Export for Executives - Create formatted Excel reports from JSON metrics.

Usage:
  python export_zcop_excel.py [--output=filename.xlsx] [--days=7]

Generates:
  - Executive Summary sheet with KPIs and trends
  - Daily Trends sheet with day-by-day breakdown
  - Dimension Analysis sheet (DMs, Business Lines, Client VPs)
  - Raw Data sheet for detailed review
  - Professional formatting with colors and charts
"""

import json
import sys
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.chart import LineChart, BarChart, PieChart, Reference
except ImportError:
    print("ERROR: openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)


class ZCOPExcelExporter:
    """Export ZCOP metrics to formatted Excel workbook."""
    
    # Color scheme
    COLORS = {
        'header': '1F4E78',      # Dark blue
        'summary': 'D9E1F2',      # Light blue
        'positive': 'C6EFCE',     # Light green
        'negative': 'FFC7CE',     # Light red
        'neutral': 'FFF2CC',      # Light yellow
        'accent': '4472C4'        # Medium blue
    }
    
    def __init__(self, metrics_file: str = None, output_file: str = None, days: int = 7):
        if not metrics_file:
            metrics_file = Path(__file__).parent.parent / 'data' / 'zcop-metrics.json'
        
        if not output_file:
            # Save to Zcop Output folder
            output_dir = Path(__file__).parent.parent.parent / 'Zcop Output'
            output_dir.mkdir(exist_ok=True)
            output_file = output_dir / f'zcop_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        self.metrics_file = Path(metrics_file)
        self.output_file = Path(output_file)
        self.days = days
        self.snapshots = []
        self.wb = None
        self.recent_snapshots = []
        
        self.load_snapshots()
    
    def load_snapshots(self) -> bool:
        """Load snapshots from metrics file."""
        try:
            if not self.metrics_file.exists():
                print(f"ERROR: Metrics file not found: {self.metrics_file}")
                return False
            
            with open(self.metrics_file, 'r') as f:
                data = json.load(f)
                self.snapshots = data.get('snapshots', [])
                self.snapshots.sort(key=lambda x: x['date'])
            
            # Get recent snapshots
            cutoff_date = (datetime.now() - timedelta(days=self.days)).strftime('%Y-%m-%d')
            self.recent_snapshots = [s for s in self.snapshots if s['date'] >= cutoff_date]
            
            return len(self.snapshots) > 0
        except Exception as e:
            print(f"ERROR: Failed to load metrics: {e}")
            return False
    
    def create_workbook(self) -> bool:
        """Create workbook with all sheets."""
        try:
            self.wb = Workbook()
            self.wb.remove(self.wb.active)  # Remove default sheet
            
            # Create sheets in order
            self.create_executive_summary()
            self.create_daily_trends()
            self.create_dimension_analysis()
            self.create_raw_data()
            
            self.wb.save(self.output_file)
            print(f"✓ Excel report saved: {self.output_file}")
            return True
        except Exception as e:
            print(f"ERROR: Failed to create workbook: {e}")
            return False
    
    def get_header_fill(self) -> PatternFill:
        """Get header fill style."""
        return PatternFill(start_color=self.COLORS['header'], end_color=self.COLORS['header'], fill_type='solid')
    
    def get_header_font(self) -> Font:
        """Get header font style."""
        return Font(bold=True, color='FFFFFF', size=11)
    
    def get_border(self) -> Border:
        """Get cell border style."""
        thin_border = Side(style='thin', color='000000')
        return Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def format_header_row(self, ws, row_num: int, columns: List[str]):
        """Format a header row."""
        for col_num, column_title in enumerate(columns, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.fill = self.get_header_fill()
            cell.font = self.get_header_font()
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.get_border()
    
    def create_executive_summary(self):
        """Create executive summary sheet."""
        ws = self.wb.create_sheet('Executive Summary', 0)
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        # Title
        ws['A1'] = 'ZCOP Analysis - Executive Summary'
        ws['A1'].font = Font(bold=True, size=14, color=self.COLORS['header'])
        ws.merge_cells('A1:D1')
        
        # Report date
        ws['A2'] = f"As of: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:D2')
        
        row = 4
        
        # Overall metrics
        ws[f'A{row}'] = 'KEY PERFORMANCE INDICATORS'
        ws[f'A{row}'].font = Font(bold=True, size=12, color=self.COLORS['header'])
        ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS['summary'], end_color=self.COLORS['summary'], fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        if self.recent_snapshots:
            total_billable = sum(s['metrics']['billable']['hours'] for s in self.recent_snapshots)
            total_non_billable = sum(s['metrics']['non_billable']['hours'] for s in self.recent_snapshots)
            total_hours = total_billable + total_non_billable
            billable_pct = (total_billable / total_hours * 100) if total_hours > 0 else 0
            
            total_nga = sum(s['metrics']['nga']['allocation_count'] for s in self.recent_snapshots)
            
            latest = self.recent_snapshots[-1]['metrics']
            prev_latest = self.recent_snapshots[-2]['metrics'] if len(self.recent_snapshots) > 1 else None
            
            # Metric rows
            metrics = [
                ('Total Billable Hours', f"{total_billable:.1f}", self.get_trend_direction(total_billable, prev_latest['billable']['hours'] if prev_latest else total_billable)),
                ('Total Non-Billable Hours', f"{total_non_billable:.1f}", self.get_trend_direction(total_non_billable, prev_latest['non_billable']['hours'] if prev_latest else total_non_billable)),
                ('Billable %', f"{billable_pct:.1f}%", 'N/A'),
                ('NGA Allocations', f"{total_nga}", 'N/A'),
                ('Ramp Up (RU) Count', f"{sum(s['metrics']['ramp_up']['count'] for s in self.recent_snapshots)}", 'Increasing resources'),
                ('Ramp Down (RD) Count', f"{sum(s['metrics']['ramp_down']['count'] for s in self.recent_snapshots)}", 'Decreasing resources'),
                ('Snapshots Collected', f"{len(self.recent_snapshots)}", 'N/A'),
            ]
            
            for metric_name, value, trend in metrics:
                ws[f'A{row}'] = metric_name
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = value
                ws[f'B{row}'].alignment = Alignment(horizontal='right')
                ws[f'C{row}'] = trend
                ws[f'C{row}'].alignment = Alignment(horizontal='center')
                
                # Color code trend
                if '↑' in trend:
                    ws[f'C{row}'].fill = PatternFill(start_color=self.COLORS['positive'], end_color=self.COLORS['positive'], fill_type='solid')
                elif '↓' in trend:
                    ws[f'C{row}'].fill = PatternFill(start_color=self.COLORS['negative'], end_color=self.COLORS['negative'], fill_type='solid')
                
                row += 1
        
        row += 1
        
        # Top dimensions
        ws[f'A{row}'] = 'TOP DIMENSIONS'
        ws[f'A{row}'].font = Font(bold=True, size=12, color=self.COLORS['header'])
        ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS['summary'], end_color=self.COLORS['summary'], fill_type='solid')
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        if self.recent_snapshots:
            # Aggregate dimension data
            dm_counts = defaultdict(lambda: {'count': 0, 'billable': 0})
            bl_counts = defaultdict(lambda: {'count': 0})
            cv_counts = defaultdict(lambda: {'count': 0})
            
            for snapshot in self.recent_snapshots:
                for dm, stats in snapshot['dimensions'].get('by_dm', {}).items():
                    dm_counts[dm]['count'] += stats.get('count', 0)
                    dm_counts[dm]['billable'] += stats.get('billable_hours', 0)
                for bl, stats in snapshot['dimensions'].get('by_business_line', {}).items():
                    bl_counts[bl]['count'] += stats.get('count', 0)
                for cv, stats in snapshot['dimensions'].get('by_client_vp', {}).items():
                    cv_counts[cv]['count'] += stats.get('count', 0)
            
            # Top 3 each
            ws[f'A{row}'] = 'Top Delivery Managers'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = 'Count'
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'C{row}'] = 'Billable Hrs'
            ws[f'C{row}'].font = Font(bold=True)
            row += 1
            
            for dm, stats in sorted(dm_counts.items(), key=lambda x: x[1]['count'], reverse=True)[:3]:
                ws[f'A{row}'] = dm
                ws[f'B{row}'] = stats['count']
                ws[f'C{row}'] = f"{stats['billable']:.1f}"
                row += 1
            
            row += 1
            
            ws[f'A{row}'] = 'Top Business Lines'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = 'Count'
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for bl, stats in sorted(bl_counts.items(), key=lambda x: x[1]['count'], reverse=True)[:3]:
                ws[f'A{row}'] = bl
                ws[f'B{row}'] = stats['count']
                row += 1
            
            row += 1
            
            ws[f'A{row}'] = 'Top Client VPs'
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = 'Count'
            ws[f'B{row}'].font = Font(bold=True)
            row += 1
            
            for cv, stats in sorted(cv_counts.items(), key=lambda x: x[1]['count'], reverse=True)[:3]:
                ws[f'A{row}'] = cv
                ws[f'B{row}'] = stats['count']
                row += 1
    
    def get_trend_direction(self, current: float, previous: float) -> str:
        """Get trend direction indicator."""
        if current == previous:
            return '→ Stable'
        elif current > previous:
            pct = ((current - previous) / previous * 100) if previous != 0 else 0
            return f'↑ +{pct:.1f}%'
        else:
            pct = ((previous - current) / previous * 100) if previous != 0 else 0
            return f'↓ -{pct:.1f}%'
    
    def create_daily_trends(self):
        """Create daily trends sheet."""
        ws = self.wb.create_sheet('Daily Trends', 1)
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 16
        
        # Header
        headers = ['Date', 'Billable Hrs', 'Non-Bill Hrs', 'Change %', 'NGA Count', 'RU Count', 'RD Count', 'File']
        self.format_header_row(ws, 1, headers)
        
        row = 2
        
        if self.recent_snapshots:
            for snapshot in self.recent_snapshots:
                m = snapshot['metrics']
                
                ws[f'A{row}'] = snapshot['date']
                ws[f'B{row}'] = m['billable']['hours']
                ws[f'B{row}'].number_format = '0.0'
                ws[f'C{row}'] = m['non_billable']['hours']
                ws[f'C{row}'].number_format = '0.0'
                
                # Trend calculation
                idx = self.recent_snapshots.index(snapshot)
                if idx > 0:
                    prev_billable = self.recent_snapshots[idx-1]['metrics']['billable']['hours']
                    change_pct = ((m['billable']['hours'] - prev_billable) / prev_billable * 100) if prev_billable != 0 else 0
                    ws[f'D{row}'] = change_pct / 100
                    ws[f'D{row}'].number_format = '0.0%'
                    
                    # Color code
                    if change_pct > 0:
                        ws[f'D{row}'].fill = PatternFill(start_color=self.COLORS['positive'], end_color=self.COLORS['positive'], fill_type='solid')
                    elif change_pct < 0:
                        ws[f'D{row}'].fill = PatternFill(start_color=self.COLORS['negative'], end_color=self.COLORS['negative'], fill_type='solid')
                else:
                    ws[f'D{row}'] = 'Baseline'
                
                ws[f'E{row}'] = m['nga']['allocation_count']
                ws[f'F{row}'] = m['ramp_up']['count']
                ws[f'G{row}'] = m['ramp_down']['count']
                ws[f'H{row}'] = snapshot['file_source']
                ws[f'H{row}'].font = Font(size=9)
                
                # Borders
                for col in range(1, 9):
                    ws.cell(row=row, column=col).border = self.get_border()
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                row += 1
    
    def create_dimension_analysis(self):
        """Create dimension analysis sheet."""
        ws = self.wb.create_sheet('Dimension Analysis', 2)
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 12
        
        row = 1
        
        if not self.recent_snapshots:
            ws[f'A{row}'] = 'No data available'
            return
        
        # Aggregate dimension data
        dm_counts = defaultdict(lambda: {'count': 0, 'billable': 0, 'appearances': 0})
        bl_counts = defaultdict(lambda: {'count': 0, 'billable': 0, 'appearances': 0})
        cv_counts = defaultdict(lambda: {'count': 0, 'billable': 0, 'appearances': 0})
        
        for snapshot in self.recent_snapshots:
            for dm, stats in snapshot['dimensions'].get('by_dm', {}).items():
                dm_counts[dm]['count'] += stats.get('count', 0)
                dm_counts[dm]['billable'] += stats.get('billable_hours', 0)
                dm_counts[dm]['appearances'] += 1
            for bl, stats in snapshot['dimensions'].get('by_business_line', {}).items():
                bl_counts[bl]['count'] += stats.get('count', 0)
                bl_counts[bl]['billable'] += stats.get('billable_hours', 0)
                bl_counts[bl]['appearances'] += 1
            for cv, stats in snapshot['dimensions'].get('by_client_vp', {}).items():
                cv_counts[cv]['count'] += stats.get('count', 0)
                cv_counts[cv]['billable'] += stats.get('billable_hours', 0)
                cv_counts[cv]['appearances'] += 1
        
        # DELIVERY MANAGERS
        ws[f'A{row}'] = 'DELIVERY MANAGERS'
        ws[f'A{row}'].font = Font(bold=True, size=12, color=self.COLORS['header'])
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        headers = ['DM Name', 'People', 'Billable Hrs', 'Appearances']
        self.format_header_row(ws, row, headers)
        row += 1
        
        for dm, stats in sorted(dm_counts.items(), key=lambda x: x[1]['count'], reverse=True):
            ws[f'A{row}'] = dm
            ws[f'B{row}'] = stats['count']
            ws[f'C{row}'] = stats['billable']
            ws[f'C{row}'].number_format = '0.0'
            ws[f'D{row}'] = stats['appearances']
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.get_border()
            
            row += 1
        
        row += 2
        
        # BUSINESS LINES
        ws[f'A{row}'] = 'BUSINESS LINES'
        ws[f'A{row}'].font = Font(bold=True, size=12, color=self.COLORS['header'])
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        headers = ['Business Line', 'People', 'Billable Hrs', 'Appearances']
        self.format_header_row(ws, row, headers)
        row += 1
        
        for bl, stats in sorted(bl_counts.items(), key=lambda x: x[1]['count'], reverse=True):
            ws[f'A{row}'] = bl
            ws[f'B{row}'] = stats['count']
            ws[f'C{row}'] = stats['billable']
            ws[f'C{row}'].number_format = '0.0'
            ws[f'D{row}'] = stats['appearances']
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.get_border()
            
            row += 1
        
        row += 2
        
        # CLIENT VPs
        ws[f'A{row}'] = 'CLIENT VPs'
        ws[f'A{row}'].font = Font(bold=True, size=12, color=self.COLORS['header'])
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        headers = ['Client VP', 'People', 'Billable Hrs', 'Appearances']
        self.format_header_row(ws, row, headers)
        row += 1
        
        for cv, stats in sorted(cv_counts.items(), key=lambda x: x[1]['count'], reverse=True):
            ws[f'A{row}'] = cv
            ws[f'B{row}'] = stats['count']
            ws[f'C{row}'] = stats['billable']
            ws[f'C{row}'].number_format = '0.0'
            ws[f'D{row}'] = stats['appearances']
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.get_border()
            
            row += 1
    
    def create_raw_data(self):
        """Create raw data sheet for detailed review."""
        ws = self.wb.create_sheet('Raw Data', 3)
        
        headers = ['Date', 'File', 'Billable Hrs', 'Non-Bill Hrs', 'Billable Cost', 'Non-Bill Cost', 'NGA Count', 'NGA %', 'RU Count', 'RD Count', 'Anomalies']
        
        col_widths = [12, 30, 14, 14, 14, 14, 12, 10, 12, 12, 40]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        self.format_header_row(ws, 1, headers)
        
        row = 2
        
        if self.recent_snapshots:
            for snapshot in self.recent_snapshots:
                m = snapshot['metrics']
                
                ws[f'A{row}'] = snapshot['date']
                ws[f'B{row}'] = snapshot['file_source']
                ws[f'C{row}'] = m['billable']['hours']
                ws[f'C{row}'].number_format = '0.0'
                ws[f'D{row}'] = m['non_billable']['hours']
                ws[f'D{row}'].number_format = '0.0'
                ws[f'E{row}'] = m['billable'].get('cost', 0)
                ws[f'E{row}'].number_format = '#,##0.00'
                ws[f'F{row}'] = m['non_billable'].get('cost', 0)
                ws[f'F{row}'].number_format = '#,##0.00'
                ws[f'G{row}'] = m['nga']['allocation_count']
                ws[f'H{row}'] = m['nga']['percentage_of_total']
                ws[f'H{row}'].number_format = '0.0%'
                ws[f'I{row}'] = m['ramp_up']['count']
                ws[f'J{row}'] = m['ramp_down']['count']
                
                # Anomalies
                anomalies = '; '.join(snapshot.get('anomalies', []))
                if anomalies:
                    ws[f'K{row}'] = anomalies
                    ws[f'K{row}'].font = Font(color='FF0000')
                
                # Borders and alignment
                for col in range(1, 12):
                    cell = ws.cell(row=row, column=col)
                    cell.border = self.get_border()
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                row += 1


def main():
    output_file = None
    days = 7
    
    # Parse arguments
    for arg in sys.argv[1:]:
        if arg.startswith('--output='):
            output_file = arg.split('=')[1]
        elif arg.startswith('--days='):
            try:
                days = int(arg.split('=')[1])
            except:
                pass
    
    exporter = ZCOPExcelExporter(output_file=output_file, days=days)
    
    if not exporter.recent_snapshots:
        print("No snapshots available. Run the ZCOP Analysis Agent first to collect data.")
        sys.exit(1)
    
    if exporter.create_workbook():
        print(f"\nReport covers {len(exporter.recent_snapshots)} days of data")
        print(f"Ready to share with executive team!")
        sys.exit(0)
    else:
        sys.exit(1)


if __name__ == '__main__':
    main()
