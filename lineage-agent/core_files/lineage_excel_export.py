# lineage_excel_export.py
# ────────────────────────────────────────────────────────────────────────────
# Builds a multi-sheet Excel workbook for a field-level lineage query.
#
# Template (4 sheets):
#   1. Summary          — target field metadata + run info
#   2. Lineage Path     — one row per edge, colour-coded by source layer
#   3. Transformation Logic — per-edge Cosmos DB summary (expressions, SQL, filters)
#   4. Transformation Chain — step-by-step detail from transformation_chain[]
#
# Layer colour scheme (matches the TiDy green palette):
#   TPR  — light blue   #D6E4F0   (source / transactional)
#   TT   — light amber  #FFF3CD   (staging / temp)
#   DDM  — light green  #D4EDDA   (data mart / reporting)
#   Header rows — dark green #1B5E20  with white text
# ────────────────────────────────────────────────────────────────────────────

import io
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ─── Colour palette ─────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1B5E20")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)

_LAYER_FILL = {
    "TPR": PatternFill("solid", fgColor="D6E4F0"),
    "TT":  PatternFill("solid", fgColor="FFF3CD"),
    "DDM": PatternFill("solid", fgColor="D4EDDA"),
}
_DEFAULT_FILL = PatternFill("solid", fgColor="F5F5F5")

_SECTION_FILL = PatternFill("solid", fgColor="E8F5E9")   # summary label rows
_SECTION_FONT = Font(bold=True, name="Calibri", size=10, color="1B5E20")

_BODY_FONT    = Font(name="Calibri", size=10)
_WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="top")

_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


# ─── Helpers ────────────────────────────────────────────────────────────────

def _header_row(ws, col_headers: list[str]):
    """Append a styled header row and return the row number."""
    ws.append(col_headers)
    row = ws.max_row
    for col_idx in range(1, len(col_headers) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.border = _BORDER
        cell.alignment = _CENTER_ALIGN
    return row


def _data_row(ws, values: list, layer: str | None = None):
    """Append a data row with optional layer colour and borders."""
    ws.append(values)
    row = ws.max_row
    fill = _LAYER_FILL.get(layer, _DEFAULT_FILL) if layer else _DEFAULT_FILL
    for col_idx in range(1, len(values) + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill      = fill
        cell.font      = _BODY_FONT
        cell.border    = _BORDER
        cell.alignment = _WRAP_ALIGN
    return row


def _set_col_widths(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ─── Sheet 1: Summary ───────────────────────────────────────────────────────

def _build_summary(wb, field_id: str, edges: list, cosmos_records: list):
    ws = wb.active
    ws.title = "Summary"

    # Derive metadata from edges
    parts      = field_id.upper().split(".")
    schema     = parts[0] if len(parts) == 3 else ""
    table      = parts[1] if len(parts) == 3 else (parts[0] if len(parts) == 2 else "")
    field      = parts[2] if len(parts) == 3 else (parts[1] if len(parts) == 2 else field_id)

    target_edges = [e for e in edges if e.get("to_field", "").upper() == field]
    data_type = ""
    if target_edges:
        data_type = target_edges[0].get("to_data_type") or ""

    layers  = sorted({e["from_layer"] for e in edges} | {e["to_layer"] for e in edges},
                     key=lambda x: {"TPR": 0, "TT": 1, "DDM": 2}.get(x, 9))
    layer_flow = " → ".join(layers)

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value     = "TiDy — Field Lineage Export"
    title_cell.fill      = _HEADER_FILL
    title_cell.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=13)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    def _kv(label, value):
        ws.append(["", ""])   # blank spacer already added outside; re-use append
        row = ws.max_row
        ws.cell(row=row, column=1).value     = label
        ws.cell(row=row, column=1).fill      = _SECTION_FILL
        ws.cell(row=row, column=1).font      = _SECTION_FONT
        ws.cell(row=row, column=1).border    = _BORDER
        ws.cell(row=row, column=1).alignment = _WRAP_ALIGN
        ws.cell(row=row, column=2).value     = str(value) if value is not None else ""
        ws.cell(row=row, column=2).font      = _BODY_FONT
        ws.cell(row=row, column=2).border    = _BORDER
        ws.cell(row=row, column=2).alignment = _WRAP_ALIGN

    # Blank row between title and data
    ws.append(["", ""])

    rows_data = [
        ("Target Field ID",   field_id.upper()),
        ("Schema",            schema),
        ("Table",             table),
        ("Field Name",        field),
        ("Data Type",         data_type),
        ("Layer Flow",        layer_flow),
        ("Layers Traversed",  ", ".join(layers)),
        ("Total Edges",       len(edges)),
        ("Cosmos Records",    len(cosmos_records)),
        ("Generated On",      datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]

    # Write rows using direct append to avoid off-by-one from _kv
    for i, (label, value) in enumerate(rows_data):
        row_num = 3 + i
        c1 = ws.cell(row=row_num, column=1, value=label)
        c1.fill = _SECTION_FILL; c1.font = _SECTION_FONT; c1.border = _BORDER; c1.alignment = _WRAP_ALIGN
        c2 = ws.cell(row=row_num, column=2, value=str(value) if value is not None else "")
        c2.font = _BODY_FONT; c2.border = _BORDER; c2.alignment = _WRAP_ALIGN

    # Layer legend
    legend_start = 3 + len(rows_data) + 2
    ws.cell(row=legend_start, column=1, value="Layer Colour Key").font = _SECTION_FONT
    ws.cell(row=legend_start, column=1).fill = _SECTION_FILL
    for offset, (lyr, fill) in enumerate(_LAYER_FILL.items(), start=1):
        r = legend_start + offset
        c = ws.cell(row=r, column=1, value=lyr)
        c.fill = fill; c.font = _BODY_FONT; c.border = _BORDER; c.alignment = _CENTER_ALIGN
        desc = {"TPR": "Source / Transactional", "TT": "Staging / Transformation", "DDM": "Data Mart / Reporting"}
        d = ws.cell(row=r, column=2, value=desc.get(lyr, ""))
        d.fill = fill; d.font = _BODY_FONT; d.border = _BORDER

    _set_col_widths(ws, [28, 60])


# ─── Sheet 2: Lineage Path ──────────────────────────────────────────────────

def _build_lineage_path(wb, edges: list):
    ws = wb.create_sheet("Lineage Path")

    headers = [
        "Hop #",
        "Source Layer", "Source Schema", "Source Table", "Source Field", "Source Data Type",
        "Target Layer", "Target Schema", "Target Table", "Target Field", "Target Data Type",
        "Mapping Name", "Transformation Name", "Transformation Type", "Expression",
    ]
    _header_row(ws, headers)
    _freeze(ws)

    # Sort edges: by from_layer order, then from_table, then to_layer
    _layer_order = {"TPR": 0, "TT": 1, "DDM": 2}
    sorted_edges = sorted(
        edges,
        key=lambda e: (
            _layer_order.get(e.get("from_layer", ""), 9),
            e.get("from_table", ""),
            _layer_order.get(e.get("to_layer", ""), 9),
        )
    )

    for hop, e in enumerate(sorted_edges, start=1):
        _data_row(ws, [
            hop,
            e.get("from_layer", ""),
            e.get("from_schema", ""),
            e.get("from_table", ""),
            e.get("from_field", ""),
            e.get("from_data_type", ""),
            e.get("to_layer", ""),
            e.get("to_schema", ""),
            e.get("to_table", ""),
            e.get("to_field", ""),
            e.get("to_data_type", ""),
            e.get("mapping_name", ""),
            e.get("transformation_name", ""),
            e.get("transformation_type", ""),
            e.get("expression", ""),
        ], layer=e.get("from_layer"))

    _set_col_widths(ws, [6, 8, 16, 22, 26, 14, 8, 16, 22, 26, 14, 40, 32, 22, 60])
    ws.row_dimensions[1].height = 20


# ─── Sheet 3: Transformation Logic ──────────────────────────────────────────

def _build_transformation_logic(wb, cosmos_records: list):
    ws = wb.create_sheet("Transformation Logic")

    headers = [
        "From Field (ID)", "To Field (ID)", "Mapping Name", "Folder",
        "Final Expression", "Custom SQL",
        "Lookup Condition", "Filter Condition",
        "Update Strategy", "Steps",
    ]
    _header_row(ws, headers)
    _freeze(ws)

    for rec in cosmos_records:
        # Infer layer for colour from to_vertex
        layer = _guess_layer(rec.get("to_vertex", ""))
        _data_row(ws, [
            rec.get("from_vertex", ""),
            rec.get("to_vertex", ""),
            rec.get("mapping_name", ""),
            rec.get("folder_name", ""),
            rec.get("final_expression", ""),
            rec.get("custom_sql", ""),
            rec.get("lookup_condition", ""),
            rec.get("filter_condition", ""),
            rec.get("update_strategy_expression", ""),
            rec.get("transformation_steps_count", ""),
        ], layer=layer)

    _set_col_widths(ws, [40, 40, 40, 22, 60, 60, 50, 50, 30, 8])
    ws.row_dimensions[1].height = 20


# ─── Sheet 4: Transformation Chain ──────────────────────────────────────────

def _build_transformation_chain(wb, cosmos_records: list):
    ws = wb.create_sheet("Transformation Chain")

    headers = [
        "From Field (ID)", "To Field (ID)", "Mapping Name",
        "Step #", "Transformation Name", "Transformation Type",
        "Input Port", "Output Port", "Expression",
    ]
    _header_row(ws, headers)
    _freeze(ws)

    for rec in cosmos_records:
        from_v   = rec.get("from_vertex", "")
        to_v     = rec.get("to_vertex", "")
        mapping  = rec.get("mapping_name", "")
        chain    = rec.get("transformation_chain", [])
        layer    = _guess_layer(to_v)

        if not chain:
            # No chain detail — write a single summary row
            _data_row(ws, [from_v, to_v, mapping, "", "", "", "", "",
                           rec.get("final_expression", "")], layer=layer)
            continue

        for step in chain:
            _data_row(ws, [
                from_v,
                to_v,
                mapping,
                step.get("step", ""),
                step.get("transformation_name", ""),
                step.get("transformation_type", ""),
                step.get("input_port", ""),
                step.get("output_port", ""),
                step.get("expression", ""),
            ], layer=layer)

    _set_col_widths(ws, [40, 40, 40, 8, 32, 22, 24, 24, 60])
    ws.row_dimensions[1].height = 20


# ─── Layer guesser from vertex ID ───────────────────────────────────────────

def _guess_layer(vertex_id: str) -> str | None:
    """Infer layer from SCHEMA.TABLE.FIELD vertex id (e.g. CRDM_DDM → DDM)."""
    upper = vertex_id.upper()
    if "_DDM" in upper or upper.startswith("DDM"):
        return "DDM"
    if "_TMP" in upper or "_TT" in upper or upper.startswith("TT"):
        return "TT"
    if "_TPR" in upper or upper.startswith("TPR"):
        return "TPR"
    return None


# ─── Public entry point ──────────────────────────────────────────────────────

def build_lineage_excel(field_id: str, edges_json: str, cosmos_json: str) -> bytes:
    """
    Build a styled multi-sheet Excel workbook and return its content as bytes.

    :param field_id:     Target field in SCHEMA.TABLE.FIELD format.
    :param edges_json:   JSON string — list of edge dicts from query_column_lineage.
    :param cosmos_json:  JSON string — list of Cosmos records from get_field_transformation_logic.
    :return:             Raw .xlsx bytes ready to send as a file download.
    """
    edges           = json.loads(edges_json) if isinstance(edges_json, str) else edges_json
    cosmos_records  = json.loads(cosmos_json) if isinstance(cosmos_json, str) else cosmos_json

    wb = Workbook()

    _build_summary(wb, field_id, edges, cosmos_records)
    _build_lineage_path(wb, edges)
    _build_transformation_logic(wb, cosmos_records)
    _build_transformation_chain(wb, cosmos_records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
