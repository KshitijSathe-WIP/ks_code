import os
import csv
import shutil
from openpyxl import load_workbook
from pathlib import Path

# Set up paths
zcop_folder = Path(r"c:\Data_KS\OneDrive - Wipro\Project Data\KS_Code\Zcop Analysis")
zcop_output_folder = Path(r"c:\Data_KS\OneDrive - Wipro\Project Data\KS_Code\Zcop Output")
temp_folder = Path(r"c:\Data_KS\OneDrive - Wipro\Project Data\KS_Code\temp")
temp_folder.mkdir(exist_ok=True)

excel_file = zcop_folder / "TD Bank ZCOP 14-Apr-2026.xlsx"
print(f"Processing: {excel_file.name}")

# Try to copy the file to temp location to avoid permission issues
temp_file = temp_folder / excel_file.name
try:
    shutil.copy2(excel_file, temp_file)
    print(f"Copied to temp: {temp_file}")
    file_to_read = temp_file
except Exception as e:
    print(f"Copy failed: {e}")
    file_to_read = excel_file

# Try reading with read_only mode
try:
    wb = load_workbook(file_to_read, data_only=True, read_only=False)
    print(f"Sheets: {wb.sheetnames}")
    
    data_rows = []
    ru_rd_rows = []
    
    # Process DATA sheet
    if "DATA" in wb.sheetnames:
        ws = wb["DATA"]
        print(f"Reading DATA sheet (max_row: {ws.max_row})")
        
        # Get header from first row
        header = []
        for cell in ws[1]:
            header.append(cell.value)
        data_rows.append(header)
        
        # Get data rows (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if any(row):  # Skip empty rows
                data_rows.append(row)
        print(f"Added {len(data_rows) - 1} rows from DATA sheet")
    
    # Process RU-RD sheet
    if "RU-RD" in wb.sheetnames:
        ws = wb["RU-RD"]
        print(f"Reading RU-RD sheet (max_row: {ws.max_row})")
        
        # Get header from first row
        header = []
        for cell in ws[1]:
            header.append(cell.value)
        ru_rd_rows.append(header)
        
        # Get data rows (skip header)
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if any(row):  # Skip empty rows
                ru_rd_rows.append(row)
        print(f"Added {len(ru_rd_rows) - 1} rows from RU-RD sheet")
    
    wb.close()
    
    # Write DATA CSV
    if data_rows:
        data_csv_path = zcop_output_folder / "DATA_14Apr_2026.csv"
        with open(data_csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(data_rows)
        print(f"\n✓ Created DATA CSV: {data_csv_path.name}")
        print(f"  Records: {len(data_rows) - 1}")
    
    # Write RU-RD CSV
    if ru_rd_rows:
        ru_rd_csv_path = zcop_output_folder / "RU-RD_14Apr_2026.csv"
        with open(ru_rd_csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(ru_rd_rows)
        print(f"✓ Created RU-RD CSV: {ru_rd_csv_path.name}")
        print(f"  Records: {len(ru_rd_rows) - 1}")
    
    print("\n✓ Complete!")

except Exception as e:
    print(f"Error: {e}")
    import traceback
    traceback.print_exc()

finally:
    # Clean up temp file
    if temp_file.exists():
        try:
            os.remove(temp_file)
            print(f"\nCleaned up temp file")
        except:
            pass
