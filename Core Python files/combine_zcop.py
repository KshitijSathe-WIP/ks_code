import os
import csv
from openpyxl import load_workbook
from pathlib import Path

# Set up paths
_BASE = Path(__file__).resolve().parent.parent
zcop_folder = _BASE / "Zcop Analysis"
zcop_output_folder = _BASE / "Zcop Output"

# Create output folder if it doesn't exist
zcop_output_folder.mkdir(exist_ok=True)

# Get all Excel files
excel_files = sorted(zcop_folder.glob("TD Bank ZCOP *.xlsx"))
print(f"Found {len(excel_files)} Excel files")

data_rows = []
ru_rd_rows = []

# Process each Excel file
for excel_file in excel_files:
    print(f"Processing: {excel_file.name}")
    try:
        wb = load_workbook(excel_file, data_only=True)
        print(f"  Sheets: {wb.sheetnames}")
        
        # Process DATA sheet
        if "DATA" in wb.sheetnames:
            ws = wb["DATA"]
            # Get header from first row
            if not data_rows:
                header = []
                for cell in ws[1]:
                    header.append(cell.value)
                data_rows.append(header)
            
            # Get data rows (skip header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):  # Skip empty rows
                    data_rows.append(row)
            print(f"    Added {ws.max_row - 1} rows from DATA sheet")
        
        # Process RU-RD sheet
        if "RU-RD" in wb.sheetnames:
            ws = wb["RU-RD"]
            # Get header from first row
            if not ru_rd_rows:
                header = []
                for cell in ws[1]:
                    header.append(cell.value)
                ru_rd_rows.append(header)
            
            # Get data rows (skip header)
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):  # Skip empty rows
                    ru_rd_rows.append(row)
            print(f"    Added {ws.max_row - 1} rows from RU-RD sheet")
        
        wb.close()
    except Exception as e:
        print(f"  Error processing {excel_file.name}: {e}")

# Write DATA CSV
if data_rows:
    data_csv_path = zcop_output_folder / "DATA_combined.csv"
    with open(data_csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(data_rows)
    print(f"\n✓ Created DATA CSV: {data_csv_path}")
    print(f"  Records: {len(data_rows) - 1}")

# Write RU-RD CSV
if ru_rd_rows:
    ru_rd_csv_path = zcop_output_folder / "RU-RD_combined.csv"
    with open(ru_rd_csv_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerows(ru_rd_rows)
    print(f"✓ Created RU-RD CSV: {ru_rd_csv_path}")
    print(f"  Records: {len(ru_rd_rows) - 1}")

print("\n✓ Complete!")
