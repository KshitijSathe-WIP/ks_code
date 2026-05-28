"""
Usage:
    python extract_zcop.py "TD Bank ZCOP 28-Apr-2026.xlsx"

Extracts DATA and RU-RD sheets from the given ZCOP Excel file.
Each daily Excel file is the authoritative snapshot for the dates it contains,
so existing rows for those dates are replaced rather than appended.
"""

import csv
import os
import shutil
import sys
import time
import warnings
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
ZCOP_FOLDER = BASE / "Zcop Analysis"
OUTPUT_FOLDER = BASE / "Zcop Output"
TEMP_FOLDER = BASE / "temp"

DATA_CSV = OUTPUT_FOLDER / "DATA_combined.csv"
RURD_CSV = OUTPUT_FOLDER / "RU-RD_combined.csv"



# Minimum columns that must always be present
MIN_DATA_COLS = {"LOAD_DATE", "EMP_CODE", "EMP_NAME"}
MIN_RURD_COLS = {"LOAD_DATE", "EMP_CODE", "EMP_NAME"}


def _csv_header(csv_path):
    """Return the header row of an existing CSV as a set, or empty set if file missing."""
    if not Path(csv_path).exists():
        return set()
    with open(csv_path, encoding="utf-8", newline="") as f:
        return set(next(csv.reader(f)))


def validate_format(wb, filename):
    """
    Validates sheet structure by checking that every column already tracked in
    the combined CSVs is present in the incoming Excel sheet (order-independent).
    Falls back to the minimum required set when no combined CSV exists yet.
    Missing columns are reported as warnings (extraction continues); missing
    sheets are hard errors (extraction is skipped).
    Returns (True, warnings_text) or (False, error_text).
    """
    errors = []
    warnings = []

    expected_data = _csv_header(DATA_CSV) or MIN_DATA_COLS
    expected_rurd = _csv_header(RURD_CSV) or MIN_RURD_COLS

    if "DATA" not in wb.sheetnames:
        errors.append("Missing sheet: DATA")
    else:
        ws = wb["DATA"]
        xl_header = {c.value for c in next(ws.iter_rows(min_row=1, max_row=1))}
        missing = {c for c in expected_data if c and c not in xl_header}
        if missing:
            warnings.append(
                f"DATA sheet is missing {len(missing)} column(s) already in combined CSV:\n"
                + "  " + ", ".join(sorted(missing))
            )

    if "RU-RD" not in wb.sheetnames:
        errors.append("Missing sheet: RU-RD")
    else:
        ws = wb["RU-RD"]
        # RU-RD has column-width metadata in row 1; real headers are in row 2
        rows = list(ws.iter_rows(min_row=1, max_row=2, values_only=True))
        header_row = rows[1] if len(rows) > 1 else rows[0]
        xl_header = {c for c in header_row if c}
        missing = {c for c in expected_rurd if c and c not in xl_header}
        if missing:
            warnings.append(
                f"RU-RD sheet is missing {len(missing)} column(s) already in combined CSV:\n"
                + "  " + ", ".join(sorted(missing))
            )

    if errors:
        return False, "\n".join(errors)
    if warnings:
        return True, "WARNINGS (columns in Excel not in combined CSV will be ignored):\n" + "\n".join(warnings)
    return True, ""


# Columns to back-fill from per-employee history when blank
DATA_FILL_COLS = ["SL_DM_ID", "SL_DM_NAME", "BILLABLE_CATEGORY"]


def fill_blank_fields(csv_path):
    """Back-fill blank SL_DM_ID, SL_DM_NAME and BILLABLE_CATEGORY using each
    employee's most-common non-blank value across all dates in the CSV."""
    with open(csv_path, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames

    present = [c for c in DATA_FILL_COLS if c in (fieldnames or [])]
    if not present:
        return 0

    # Build per-employee lookup of most-common non-blank value
    acc = {}
    for row in rows:
        code = row["EMP_CODE"]
        if code not in acc:
            acc[code] = {c: [] for c in present}
        for col in present:
            if row[col].strip():
                acc[code][col].append(row[col])

    best = {
        code: {
            col: Counter(vals[col]).most_common(1)[0][0] if vals[col] else ""
            for col in present
        }
        for code, vals in acc.items()
    }

    fixed = 0
    for row in rows:
        code = row["EMP_CODE"]
        for col in present:
            if not row[col].strip() and best[code][col]:
                row[col] = best[code][col]
                fixed += 1

    if fixed:
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)

    return fixed


def count_records(csv_path):
    with open(csv_path, encoding="utf-8") as f:
        return sum(1 for _ in csv.reader(f)) - 1


def replace_dates_in_csv(csv_path, xl_header, new_rows, dates_to_replace):
    """Remove rows for given dates, then insert new_rows remapped to CSV column order by name.
    Only columns already present in the CSV are written; extra Excel columns are ignored.
    If the CSV does not exist yet (new output folder), it is created from scratch."""
    csv_path = Path(csv_path)
    if not csv_path.exists():
        # First run into this folder — write header + all rows directly
        clean_header = [str(c) if c is not None else "" for c in xl_header]
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(clean_header)
            w.writerows(new_rows)
        return 0, len(new_rows)

    with open(csv_path, encoding="utf-8", newline="") as f:
        all_rows = list(csv.reader(f))
    csv_header = list(all_rows[0])

    xl_col_idx = {col: i for i, col in enumerate(xl_header) if col}
    csv_col_pos = {col: i for i, col in enumerate(csv_header)}

    def remap(xl_row):
        out = [""] * len(csv_header)
        for col, xi in xl_col_idx.items():
            if col in csv_col_pos:
                val = xl_row[xi] if xi < len(xl_row) else None
                out[csv_col_pos[col]] = "" if val is None else val
        return out

    remapped_new_rows = [remap(r) for r in new_rows]
    kept = [r for r in all_rows[1:] if r and str(r[0]) not in dates_to_replace and str(r[0]) != "LOAD_DATE"]

    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(csv_header)
        w.writerows(kept)
        w.writerows(remapped_new_rows)
    return len(all_rows) - 1 - len(kept), len(remapped_new_rows)


def extract(filename, copy_folder=None):
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)

    # Accept either a bare filename (looked up in ZCOP_FOLDER) or a full path
    source = Path(filename)
    if not source.is_absolute():
        source = ZCOP_FOLDER / filename
    if not source.exists():
        print(f"ERROR: File not found: {source}")
        sys.exit(1)

    TEMP_FOLDER.mkdir(exist_ok=True)
    temp = TEMP_FOLDER / source.name

    # Copy to temp — retry up to 3 times in case the file is momentarily locked
    # (e.g. OneDrive sync, antivirus scan, or Excel still releasing the handle)
    for attempt in range(1, 4):
        try:
            shutil.copy2(source, temp)
            break
        except PermissionError:
            if attempt == 3:
                print(f"ERROR: Cannot read '{source.name}' — the file is open in another")
                print(f"       application (Excel, OneDrive, etc.). Close it and try again.")
                sys.exit(1)
            print(f"  File locked, retrying in 2 s (attempt {attempt}/3)…")
            time.sleep(2)

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(
                "ignore",
                message=".*marked as a date but the serial value.*",
                category=UserWarning,
            )
            wb = load_workbook(temp, data_only=True)

        print(f"Sheets found: {wb.sheetnames}")

        valid, reason = validate_format(wb, filename)
        if not valid:
            wb.close()
            print(f"\nFORMAT VALIDATION FAILED — {filename}")
            print(reason)
            print("Extraction skipped. Please check the file format.")
            sys.exit(2)
        if reason:
            print(reason)

        data_rows, data_xl_header = [], []
        rurd_rows, rurd_xl_header = [], []

        if "DATA" in wb.sheetnames:
            ws = wb["DATA"]
            xl_rows = list(ws.iter_rows(values_only=True))
            data_xl_header = [str(c) if c is not None else "" for c in xl_rows[0]]
            for row in xl_rows[1:]:
                if any(row) and str(row[0]) != "LOAD_DATE":
                    data_rows.append(row)

        if "RU-RD" in wb.sheetnames:
            ws = wb["RU-RD"]
            xl_rows = list(ws.iter_rows(values_only=True))
            # RU-RD row 1 is metadata; real headers are in row 2
            rurd_xl_header = [str(c) if c is not None else "" for c in xl_rows[1]]
            for row in xl_rows[2:]:
                if any(row) and str(row[0]) != "LOAD_DATE":
                    rurd_rows.append(row)

        wb.close()
    finally:
        os.remove(temp)

    # Dates present in each sheet (skip any stray header rows)
    data_dates = {str(r[0]) for r in data_rows if r[0] and str(r[0]) != "LOAD_DATE"}
    rurd_dates = {str(r[0]) for r in rurd_rows if r[0] and str(r[0]) != "LOAD_DATE"}

    removed_d, added_d = replace_dates_in_csv(DATA_CSV, data_xl_header, data_rows, data_dates)
    removed_r, added_r = replace_dates_in_csv(RURD_CSV, rurd_xl_header, rurd_rows, rurd_dates)

    filled_d = fill_blank_fields(DATA_CSV)

    data_total = count_records(DATA_CSV)
    rurd_total = count_records(RURD_CSV)

    print(f"\n{'='*57}")
    print(f"ZCOP Extraction Complete — {filename}")
    print(f"{'='*57}")
    print(f"  DATA  : replaced {removed_d} old rows -> inserted {added_d} rows" + (f" | back-filled {filled_d} blank cell(s)" if filled_d else ""))
    print(f"          dates: {sorted(data_dates)}")
    print(f"  RU-RD : replaced {removed_r} old rows -> inserted {added_r} rows")
    print(f"          dates: {sorted(rurd_dates)}")
    print(f"  DATA_combined total : {data_total} records")
    print(f"  RU-RD_combined total: {rurd_total} records")
    print(f"{'='*57}")

    # --- Copy to user-provided folder (if different from primary OUTPUT_FOLDER) ---
    if copy_folder and copy_folder.resolve() != OUTPUT_FOLDER.resolve():
        copy_folder.mkdir(parents=True, exist_ok=True)
        shutil.copy2(DATA_CSV, copy_folder / "DATA_combined.csv")
        shutil.copy2(RURD_CSV, copy_folder / "RU-RD_combined.csv")
        print(f"  Copied to          : {copy_folder}")
    print(f"{'='*57}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Extract ZCOP Excel data to combined CSVs.")
    parser.add_argument("filename", help="ZCOP Excel filename (must be in the Zcop Analysis folder)")
    parser.add_argument(
        "--output", metavar="FOLDER", default=None,
        help="Override output folder for DATA_combined.csv and RU-RD_combined.csv"
    )
    args = parser.parse_args()

    copy_folder = Path(args.output) if args.output else None
    extract(args.filename, copy_folder)
