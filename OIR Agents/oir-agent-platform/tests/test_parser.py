"""Tests for header mapping and sheet resolution."""
import pytest
from functions.ingest_oir.header_map import build_column_map, REQUIRED_FIELDS


class TestHeaderMapping:

    def _minimal_headers(self, overrides: dict | None = None) -> list[str]:
        base = {
            "demand_id": "Demand ID",
            "project": "Project",
            "role": "Role Name",
            "status": "Status",
            "pm_name": "PM Name",
            "dem_end_date": "DEM End Date",
            "comments": "Comments",
            "remarks_status": "Remarks Status",
        }
        if overrides:
            base.update(overrides)
        return list(base.values())

    def test_standard_headers_resolve(self):
        col_map = build_column_map(self._minimal_headers())
        assert "demand_id" in col_map
        assert "project" in col_map
        assert "dem_end_date" in col_map

    def test_alias_headers_resolve(self):
        headers = ["Sr_ID", "Project Name", "Role", "Category",
                   "SL_PM_Name", "Demand End Date", "Comment", "Remarks Status"]
        col_map = build_column_map(headers)
        assert "demand_id" in col_map
        assert "project" in col_map
        assert "pm_name" in col_map
        assert "dem_end_date" in col_map
        assert "comments" in col_map

    def test_column_order_irrelevant(self):
        headers = ["Remarks Status", "Comments", "DEM End Date", "Status",
                   "PM Name", "Role Name", "Project", "Demand ID"]
        col_map = build_column_map(headers)
        assert col_map["demand_id"] == 7   # last column
        assert col_map["remarks_status"] == 0  # first column

    def test_missing_required_header_raises(self):
        headers = ["Project", "Role Name", "Status", "PM Name",
                   "DEM End Date", "Comments", "Remarks Status"]
        # demand_id is missing
        with pytest.raises(ValueError, match="demand_id"):
            build_column_map(headers)

    def test_extra_columns_are_tolerated(self):
        headers = self._minimal_headers() + ["Extra Col 1", "Extra Col 2"]
        col_map = build_column_map(headers)
        assert "demand_id" in col_map   # still maps correctly

    def test_case_insensitive_matching(self):
        headers = ["DEMAND ID", "PROJECT", "ROLE NAME", "STATUS",
                   "PM NAME", "DEM END DATE", "COMMENTS", "REMARKS STATUS"]
        col_map = build_column_map(headers)
        for field in REQUIRED_FIELDS:
            assert field in col_map


class TestSheetResolution:

    def test_or_sheet_resolved(self):
        import openpyxl, tempfile, os
        wb = openpyxl.Workbook()
        wb.active.title = "Cover"
        wb.create_sheet("OR 04-08-2026")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp = f.name
        wb.save(tmp)

        from functions.ingest_oir.parser import resolve_or_sheet
        loaded = openpyxl.load_workbook(tmp)
        sheet = resolve_or_sheet(loaded)
        assert sheet.title == "OR 04-08-2026"
        loaded.close()
        os.unlink(tmp)

    def test_latest_or_sheet_selected_when_multiple(self):
        import openpyxl, tempfile, os
        wb = openpyxl.Workbook()
        wb.active.title = "Cover"
        wb.create_sheet("OR 03-08-2026")
        wb.create_sheet("OR 04-08-2026")
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp = f.name
        wb.save(tmp)

        from functions.ingest_oir.parser import resolve_or_sheet
        loaded = openpyxl.load_workbook(tmp)
        sheet = resolve_or_sheet(loaded)
        assert sheet.title == "OR 04-08-2026"
        loaded.close()
        os.unlink(tmp)

    def test_no_or_sheet_raises(self):
        import openpyxl, tempfile, os
        from functions.ingest_oir.parser import resolve_or_sheet
        from functions.shared.models import IngestionError

        wb = openpyxl.Workbook()
        wb.active.title = "Summary"
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            tmp = f.name
        wb.save(tmp)

        loaded = openpyxl.load_workbook(tmp)
        with pytest.raises(IngestionError, match="No sheet"):
            resolve_or_sheet(loaded)
        loaded.close()
        os.unlink(tmp)
