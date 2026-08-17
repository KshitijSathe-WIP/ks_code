"""OIR Excel parser.

Resolves the correct sheet by 'OR <date>' prefix, maps headers by
normalised text (not index), and yields raw row dicts ready for hashing
and upsert. Fails loudly on any structural problem — never silently
produces nulls for required fields.

All data read from the Excel file is treated as untrusted input.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date
from typing import Generator, Optional

import openpyxl
from openpyxl.workbook import Workbook

from functions.shared.models import IngestionError
from .header_map import build_column_map

logger = logging.getLogger(__name__)


@dataclass
class RawRow:
    demand_id: str
    project: str
    sldu: str
    role: str
    skill: str
    status: str
    pm_name: str
    tm_name: str
    em_name: str
    dem_start_date: Optional[date]
    dem_end_date: Optional[date]
    comments: str
    remarks_status: str
    source_file: str
    # Parent requisition (RLS_ID). Many positions share one -- demand_id is
    # the per-position SR_ID_2. Kept so a demand can be traced back.
    requisition_id: str = ""
    # Owner emails, when the file carries them. Empty when the columns are
    # absent (the current state of the real OIR report) -- see ADR 0008.
    pm_email: str = ""
    tm_email: str = ""
    em_email: str = ""
    dm_name: str = ""
    dm_email: str = ""


def _extract_trailing_date(sheet_name: str) -> str:
    """Extract trailing date portion for sort purposes (lex-safe if zero-padded)."""
    match = re.search(r"(\d{2}[-/]\d{2}[-/]\d{4})$", sheet_name.strip())
    return match.group(1) if match else ""


def resolve_or_sheet(workbook: Workbook):
    """Find the data sheet; pick the latest if several match.

    Real files name this sheet 'OIR 6th Aug ' (note: 'OIR', a written-out
    date, and often a trailing space) -- the original 'OR <date>' rule
    matched none of them. Pivot/summary tabs in the same workbook
    ('OIR Pivot', 'AXNB Pivot') must be excluded, since they also begin
    with 'OIR'.
    """
    excluded = ("pivot", "summary")
    candidates = [
        s for s in workbook.sheetnames
        if s.strip().upper().startswith(("OIR ", "OR "))
        and not any(word in s.strip().lower() for word in excluded)
    ]
    if not candidates:
        raise IngestionError(
            f"No OIR data sheet found. Sheets present: {workbook.sheetnames}. "
            "Expected one named like 'OIR <date>' (pivot/summary tabs are ignored)."
        )
    if len(candidates) > 1:
        candidates.sort(key=_extract_trailing_date, reverse=True)
        logger.warning("Multiple OIR sheets found (%s); using '%s'", candidates, candidates[0])
    return workbook[candidates[0]]


def _cell_str(value) -> str:
    """Coerce a cell value to a stripped string."""
    if value is None:
        return ""
    return str(value).strip()


def _cell_date(value) -> Optional[date]:
    """Coerce a cell value to date; return None if unparseable."""
    if value is None:
        return None
    # datetime is a subclass of date; extract .date() to avoid leaking time component
    from datetime import datetime as _dt
    if isinstance(value, _dt):
        return value.date()
    if isinstance(value, date):
        return value
    try:
        from datetime import datetime as _dt
        return _dt.strptime(str(value).strip(), "%Y-%m-%d").date()
    except ValueError:
        try:
            return _dt.strptime(str(value).strip(), "%d/%m/%Y").date()
        except ValueError:
            logger.debug("Unparseable date value '%s'", value)
            return None


def parse_workbook(path: str, source_file: str) -> Generator[RawRow, None, None]:
    """Open the workbook, find the OR sheet, and yield one RawRow per data row.

    Skips blank rows (no demand_id). Raises IngestionError on structural
    problems. Does not validate field values — that is the caller's job.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise IngestionError(f"Cannot open workbook '{path}': {exc}") from exc

    sheet = resolve_or_sheet(wb)
    rows = list(sheet.iter_rows(values_only=True))

    if not rows:
        raise IngestionError("Sheet is empty")

    # Find header row — first row that contains a recognisable demand header
    header_row_idx = None
    col_map = None
    for i, row in enumerate(rows):
        try:
            col_map = build_column_map([str(c) if c is not None else "" for c in row])
            header_row_idx = i
            break
        except ValueError:
            continue

    if header_row_idx is None or col_map is None:
        raise IngestionError(
            "Could not locate a valid header row. "
            "Check that required columns are present."
        )

    data_rows = rows[header_row_idx + 1:]
    error_count = 0

    def _get(row, field: str) -> str:
        """Value for an optional canonical field, '' when the column is absent."""
        idx = col_map.get(field)
        if idx is None or idx >= len(row):
            return ""
        return _cell_str(row[idx])

    for row in data_rows:
        demand_id = _cell_str(row[col_map["demand_id"]])
        if not demand_id:
            continue  # skip blank rows

        try:
            skill = _get(row, "skill")
            yield RawRow(
                demand_id=demand_id,
                project=_get(row, "project"),
                sldu=_get(row, "sldu"),
                # Real files carry no role column; the essential skill is what
                # a reader actually recognises the demand by, so fall back to it.
                role=_get(row, "role") or skill,
                skill=skill,
                status=_get(row, "status"),
                pm_name=_get(row, "pm_name"),
                tm_name=_get(row, "tm_name"),
                em_name=_get(row, "em_name"),
                dm_name=_get(row, "dm_name"),
                dem_start_date=_cell_date(row[col_map["dem_start_date"]] if "dem_start_date" in col_map else None),
                dem_end_date=_cell_date(row[col_map["dem_end_date"]]),
                comments=_get(row, "comments"),
                remarks_status=_get(row, "remarks_status"),
                source_file=source_file,
                requisition_id=_get(row, "requisition_id"),
                pm_email=_get(row, "pm_email"),
                tm_email=_get(row, "tm_email"),
                em_email=_get(row, "em_email"),
                dm_email=_get(row, "dm_email"),
            )
        except (IndexError, TypeError) as exc:
            error_count += 1
            logger.warning("Row parse error for demand '%s': %s", demand_id, exc)
            continue

    wb.close()

    if error_count > 0:
        logger.warning("Total row parse errors: %d", error_count)
